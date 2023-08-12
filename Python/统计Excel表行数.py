from openpyxl import load_workbook
wb = load_workbook(filename=r'/Users/igamenine/Library/CloudStorage/OneDrive-个人/文档/Python/交接文档整理/服务器crontab脚本释义.xlsx')  ##读取路径
ws = wb.get_sheet_by_name("脚本释义")  ##读取名字为Sheet1的sheet表
num = 1

while 1:
    cell = ws.cell(row=num, column=1).value
    if cell:
        # print(cell)
        num = num +1
    else:
        print("行数为", num - 1)
        exit()